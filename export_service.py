"""
Export & Reporting Service for StructuraQS
Generates professional Excel workbooks (.xlsx) with openpyxl, CSV files, and print-ready HTML/PDF reports.
Zero external API calls.
"""

import io
import csv
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Design Palette for Excel Sheets
NAVY_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
SUBHEADER_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")  # Slate 700
ACCENT_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")     # Slate 100
TOTAL_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")      # Slate 200

FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="1E293B")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
FONT_REGULAR = Font(name="Calibri", size=10, color="1E293B")
FONT_MUTED = Font(name="Calibri", size=9, italic=True, color="64748B")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
TOTAL_BORDER = Border(
    top=Side(border_style="thin", color="1E293B"),
    bottom=Side(border_style="double", color="1E293B")
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def export_boq_to_excel(project: Dict[str, Any], boq_items: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Generates a full professional BOQ Excel workbook with:
    - Cover / Summary Tab
    - Itemized BOQ Tab with Division Groups and formula totals
    - Rate Breakdown Tab
    """
    wb = openpyxl.Workbook()
    currency = project.get('currency', '$')
    currency_fmt = f'"{currency}"#,##0.00'

    # -------------------------------------------------------------
    # TAB 1: BOQ SUMMARY
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Block
    ws_summary["A1"] = "STRUCTURA-QS | QUANTITY SURVEYOR BILL OF QUANTITIES"
    ws_summary["A1"].font = FONT_TITLE
    ws_summary["A2"] = f"PROJECT: {project.get('name', 'Construction Project')} ({project.get('code', '')})"
    ws_summary["A2"].font = Font(name="Calibri", size=11, bold=True, color="334155")
    ws_summary["A3"] = f"Client: {project.get('client', '-')} | Contractor: {project.get('contractor', '-')} | Consultant: {project.get('consultant', '-')}"
    ws_summary["A3"].font = FONT_MUTED

    headers_summary = ["Div No.", "Division Work Package Description", "No. of Items", f"Division Total ({currency})", "% of Contract"]
    ws_summary.append([]) # Row 4 empty
    ws_summary.append(headers_summary) # Row 5

    for col_idx, text in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=5, column=col_idx)
        cell.fill = NAVY_HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER if col_idx in [1, 3] else (ALIGN_RIGHT if col_idx >= 4 else ALIGN_LEFT)
        cell.border = THIN_BORDER

    # Group by division
    divisions = {}
    for item in boq_items:
        div = item.get('division', 'General')
        if div not in divisions:
            divisions[div] = {'count': 0, 'total': 0.0}
        divisions[div]['count'] += 1
        divisions[div]['total'] += float(item.get('total_amount', 0.0))

    grand_total = sum(d['total'] for d in divisions.values())

    curr_row = 6
    for div_name, data in sorted(divisions.items()):
        div_code = div_name.split('.')[0] if '.' in div_name else "00"
        pct = (data['total'] / grand_total * 100.0) if grand_total > 0 else 0.0

        ws_summary.append([
            div_code,
            div_name,
            data['count'],
            data['total'],
            pct / 100.0
        ])
        
        ws_summary.cell(row=curr_row, column=1).alignment = ALIGN_CENTER
        ws_summary.cell(row=curr_row, column=2).alignment = ALIGN_LEFT
        ws_summary.cell(row=curr_row, column=3).alignment = ALIGN_CENTER
        ws_summary.cell(row=curr_row, column=4).alignment = ALIGN_RIGHT
        ws_summary.cell(row=curr_row, column=4).number_format = currency_fmt
        ws_summary.cell(row=curr_row, column=5).alignment = ALIGN_RIGHT
        ws_summary.cell(row=curr_row, column=5).number_format = "0.00%"

        for col_idx in range(1, 6):
            cell = ws_summary.cell(row=curr_row, column=col_idx)
            cell.border = THIN_BORDER
            cell.font = FONT_REGULAR

        curr_row += 1

    # Grand Total Row
    ws_summary.append(["", "TOTAL CONTRACT BILL OF QUANTITIES (DIVISIONS 01 TO 08)", len(boq_items), grand_total, 1.0])
    for col_idx in range(1, 6):
        cell = ws_summary.cell(row=curr_row, column=col_idx)
        cell.fill = TOTAL_FILL
        cell.font = FONT_BOLD
        cell.border = TOTAL_BORDER
        if col_idx == 4:
            cell.number_format = currency_fmt
            cell.alignment = ALIGN_RIGHT
        elif col_idx == 5:
            cell.number_format = "0.00%"
            cell.alignment = ALIGN_RIGHT
        elif col_idx in [1, 3]:
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_LEFT

    # Auto adjust column widths
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -------------------------------------------------------------
    # TAB 2: DETAILED ITEM-WISE BOQ
    # -------------------------------------------------------------
    ws_detail = wb.create_sheet(title="Detailed BOQ")
    ws_detail.views.sheetView[0].showGridLines = True

    # Title
    ws_detail["A1"] = f"BILL OF QUANTITIES - {project.get('name', '')}"
    ws_detail["A1"].font = FONT_TITLE
    ws_detail["A2"] = f"Standard Measurement Code: SMM7 / NRM2 / IS 1200 | All rates inclusive of Contractor Overhead & Profit"
    ws_detail["A2"].font = FONT_MUTED

    headers_detail = ["Item Code", "Division / Work Section", "Detailed Item Description & Specification", "Unit", "Quantity", f"Unit Rate ({currency})", f"Total Amount ({currency})"]
    ws_detail.append([]) # Row 3
    ws_detail.append(headers_detail) # Row 4

    for col_idx, text in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=4, column=col_idx)
        cell.fill = NAVY_HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER if col_idx in [1, 4] else (ALIGN_RIGHT if col_idx in [5, 6, 7] else ALIGN_LEFT)
        cell.border = THIN_BORDER

    current_div = None
    row_idx = 5
    for item in boq_items:
        div = item.get('division', 'General')
        if div != current_div:
            # Add division separator banner
            current_div = div
            ws_detail.append([div, "", "", "", "", "", ""])
            for c in range(1, 8):
                cell = ws_detail.cell(row=row_idx, column=c)
                cell.fill = SUBHEADER_FILL
                cell.font = FONT_SUBHEADER
                cell.border = THIN_BORDER
            row_idx += 1

        qty = float(item.get('quantity', 0.0))
        rate = float(item.get('unit_rate', 0.0))
        amt = float(item.get('total_amount', qty * rate))

        ws_detail.append([
            item.get('item_code', ''),
            item.get('section', ''),
            item.get('description', ''),
            item.get('unit', ''),
            qty,
            rate,
            amt
        ])

        ws_detail.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
        ws_detail.cell(row=row_idx, column=2).alignment = ALIGN_LEFT
        ws_detail.cell(row=row_idx, column=3).alignment = ALIGN_LEFT
        ws_detail.cell(row=row_idx, column=4).alignment = ALIGN_CENTER
        ws_detail.cell(row=row_idx, column=5).alignment = ALIGN_RIGHT
        ws_detail.cell(row=row_idx, column=5).number_format = "#,##0.00"
        ws_detail.cell(row=row_idx, column=6).alignment = ALIGN_RIGHT
        ws_detail.cell(row=row_idx, column=6).number_format = currency_fmt
        ws_detail.cell(row=row_idx, column=7).alignment = ALIGN_RIGHT
        ws_detail.cell(row=row_idx, column=7).number_format = currency_fmt

        for c in range(1, 8):
            cell = ws_detail.cell(row=row_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = FONT_REGULAR

        row_idx += 1

    # Total Row
    ws_detail.append(["", "", "TOTAL MEASURED CARRIED TO SUMMARY", "", "", "", grand_total])
    for c in range(1, 8):
        cell = ws_detail.cell(row=row_idx, column=c)
        cell.fill = TOTAL_FILL
        cell.font = FONT_BOLD
        cell.border = TOTAL_BORDER
        if c == 7:
            cell.number_format = currency_fmt
            cell.alignment = ALIGN_RIGHT
        else:
            cell.alignment = ALIGN_LEFT

    # Column widths
    ws_detail.column_dimensions['A'].width = 12
    ws_detail.column_dimensions['B'].width = 24
    ws_detail.column_dimensions['C'].width = 50
    ws_detail.column_dimensions['D'].width = 10
    ws_detail.column_dimensions['E'].width = 14
    ws_detail.column_dimensions['F'].width = 16
    ws_detail.column_dimensions['G'].width = 18

    # -------------------------------------------------------------
    # TAB 3: RATE BREAKDOWN ANALYSIS
    # -------------------------------------------------------------
    ws_rates = wb.create_sheet(title="Rate Buildup Breakdown")
    ws_rates.views.sheetView[0].showGridLines = True

    ws_rates["A1"] = "BOQ ITEM RATE ANALYSIS COMPOSITION"
    ws_rates["A1"].font = FONT_TITLE

    headers_rates = ["Item Code", "Item Description", "Unit", "Material Rate", "Labor Rate", "Plant & Equipment", "Subcontractor", "O&P %", "Final Unit Rate"]
    ws_rates.append([])
    ws_rates.append(headers_rates)

    for col_idx, text in enumerate(headers_rates, 1):
        cell = ws_rates.cell(row=3, column=col_idx)
        cell.fill = NAVY_HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER if col_idx in [1, 3] else (ALIGN_RIGHT if col_idx >= 4 else ALIGN_LEFT)
        cell.border = THIN_BORDER

    r_idx = 4
    for item in boq_items:
        ws_rates.append([
            item.get('item_code', ''),
            item.get('description', ''),
            item.get('unit', ''),
            float(item.get('material_rate', 0.0)),
            float(item.get('labor_rate', 0.0)),
            float(item.get('equipment_rate', 0.0)),
            float(item.get('subcontractor_rate', 0.0)),
            float(item.get('overhead_profit_pct', 15.0)) / 100.0,
            float(item.get('unit_rate', 0.0))
        ])

        ws_rates.cell(row=r_idx, column=1).alignment = ALIGN_CENTER
        ws_rates.cell(row=r_idx, column=2).alignment = ALIGN_LEFT
        ws_rates.cell(row=r_idx, column=3).alignment = ALIGN_CENTER
        
        for c in range(4, 8):
            ws_rates.cell(row=r_idx, column=c).alignment = ALIGN_RIGHT
            ws_rates.cell(row=r_idx, column=c).number_format = currency_fmt
            
        ws_rates.cell(row=r_idx, column=8).alignment = ALIGN_RIGHT
        ws_rates.cell(row=r_idx, column=8).number_format = "0.0%"
        ws_rates.cell(row=r_idx, column=9).alignment = ALIGN_RIGHT
        ws_rates.cell(row=r_idx, column=9).number_format = currency_fmt

        for c in range(1, 10):
            cell = ws_rates.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = FONT_REGULAR

        r_idx += 1

    ws_rates.column_dimensions['A'].width = 12
    ws_rates.column_dimensions['B'].width = 45
    ws_rates.column_dimensions['C'].width = 10
    for l in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws_rates.column_dimensions[l].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_bbs_to_excel(project: Dict[str, Any], bbs_items: List[Dict[str, Any]]) -> io.BytesIO:
    """Generates Bar Bending Schedule (BBS) Excel workbook with diameter matrix."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bar Bending Schedule"
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = f"BAR BENDING SCHEDULE (BBS) - {project.get('name', '')}"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = "Code standard: IS 2502 / BS 8666 | Steel: TMT Fe500D High Yield Rebar"
    ws["A2"].font = FONT_MUTED

    headers = [
        "Member / Location", "Bar Mark", "Shape Type", "Dia (mm)", "Cut Length (m)",
        "No. of Members", "Bars/Member", "Total Bars", "Total Length (m)",
        "Unit Wt (kg/m)", "Total Wt (kg)", "Total Wt (MT)"
    ]
    ws.append([])
    ws.append(headers)

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = NAVY_HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER if col_idx in [2, 3, 4] else (ALIGN_RIGHT if col_idx >= 5 else ALIGN_LEFT)
        cell.border = THIN_BORDER

    tot_wt_kg = 0.0
    tot_len_m = 0.0
    dia_weights = {8: 0.0, 10: 0.0, 12: 0.0, 16: 0.0, 20: 0.0, 25: 0.0, 32: 0.0}

    row_idx = 5
    for item in bbs_items:
        dia = float(item.get('diameter_mm', 0))
        wt_kg = float(item.get('total_weight_kg', 0.0))
        len_m = float(item.get('total_length_m', 0.0))
        tot_wt_kg += wt_kg
        tot_len_m += len_m

        if int(dia) in dia_weights:
            dia_weights[int(dia)] += wt_kg
        else:
            dia_weights[int(dia)] = wt_kg

        ws.append([
            item.get('member_name', ''),
            item.get('bar_mark', ''),
            item.get('shape_code', ''),
            dia,
            float(item.get('cut_length_m', 0.0)),
            int(item.get('num_members', 1)),
            int(item.get('bars_per_member', 1)),
            int(item.get('total_bars', 1)),
            len_m,
            float(item.get('unit_weight_kg_m', 0.0)),
            wt_kg,
            wt_kg / 1000.0
        ])

        ws.cell(row=row_idx, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=2).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=3).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=4).alignment = ALIGN_CENTER
        for c in range(5, 13):
            ws.cell(row=row_idx, column=c).alignment = ALIGN_RIGHT
            ws.cell(row=row_idx, column=c).number_format = "#,##0.00" if c != 12 else "#,##0.000"

        for c in range(1, 13):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = FONT_REGULAR

        row_idx += 1

    # Total Summary
    ws.append(["TOTAL REBAR STEEL REINFORCEMENT", "", "", "", "", "", "", "", tot_len_m, "", tot_wt_kg, tot_wt_kg / 1000.0])
    for c in range(1, 13):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = TOTAL_FILL
        cell.font = FONT_BOLD
        cell.border = TOTAL_BORDER
        if c in [9, 11, 12]:
            cell.alignment = ALIGN_RIGHT
            cell.number_format = "#,##0.00" if c != 12 else "#,##0.000"

    # Diameter breakdown matrix
    row_idx += 3
    ws.cell(row=row_idx, column=1, value="DIAMETER-WISE STEEL CONSUMPTION SUMMARY").font = Font(bold=True, size=11)
    row_idx += 1
    
    dia_headers = ["Bar Diameter", "Unit Weight (kg/m)", "Total Weight (kg)", "Total Weight (MT)", "% of Total Steel"]
    for i, h in enumerate(dia_headers, 1):
        c = ws.cell(row=row_idx, column=i, value=h)
        c.fill = SUBHEADER_FILL
        c.font = FONT_SUBHEADER
        c.border = THIN_BORDER

    row_idx += 1
    for d, w in sorted(dia_weights.items()):
        if w > 0:
            unit_w = (d ** 2) / 162.2
            pct = (w / tot_wt_kg * 100.0) if tot_wt_kg > 0 else 0.0
            ws.append([f"Ø {d} mm", round(unit_w, 3), round(w, 2), round(w / 1000.0, 3), pct / 100.0])
            
            ws.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=2).alignment = ALIGN_RIGHT
            ws.cell(row=row_idx, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row_idx, column=3).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=4).alignment = ALIGN_RIGHT
            ws.cell(row=row_idx, column=4).number_format = "#,##0.000"
            ws.cell(row=row_idx, column=5).alignment = ALIGN_RIGHT
            ws.cell(row=row_idx, column=5).number_format = "0.00%"

            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).border = THIN_BORDER

            row_idx += 1

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_boq_to_csv(boq_items: List[Dict[str, Any]]) -> str:
    """Generates standard CSV representation of BOQ."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Item Code", "Division", "Section", "Description", "Unit", "Quantity", "Unit Rate", "Total Amount"])
    for item in boq_items:
        writer.writerow([
            item.get('item_code', ''),
            item.get('division', ''),
            item.get('section', ''),
            item.get('description', ''),
            item.get('unit', ''),
            item.get('quantity', 0.0),
            item.get('unit_rate', 0.0),
            item.get('total_amount', 0.0)
        ])
    return output.getvalue()
