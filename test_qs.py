"""
Comprehensive Verification Test Suite for StructuraQS
Tests pure mathematical calculation engines, SQLite queries, REST endpoints, and Excel export generation.
"""

import unittest
import json
import io
import openpyxl

import db
import calculators
import export_service
from app import app

class TestStructuraQS(unittest.TestCase):
    def setUp(self):
        self.app = app.test_client()
        self.app.testing = True
        db.init_db()

    def test_bbs_all_shape_codes(self):
        # 1. Straight
        res1 = calculators.calculate_bbs_item('STRAIGHT', 16.0, {'a': 6.0}, 10, 4)
        self.assertEqual(res1['total_bars'], 40)
        self.assertEqual(res1['cut_length_m'], 6.0)
        self.assertAlmostEqual(res1['unit_weight_kg_m'], 1.578, delta=0.01)

        # 2. L-Bend (90 deg)
        res2 = calculators.calculate_bbs_item('L_BEND', 20.0, {'a': 4.0, 'b': 0.5}, 2, 6)
        self.assertAlmostEqual(res2['cut_length_m'], 4.46, places=2)

        # 3. U-Shape (180 deg)
        res3 = calculators.calculate_bbs_item('U_SHAPE', 12.0, {'a': 3.0, 'b': 0.4}, 4, 2)
        # 3.0 + 2*0.4 - 4*0.012 = 3.8 - 0.048 = 3.752m
        self.assertAlmostEqual(res3['cut_length_m'], 3.752, places=3)

        # 4. Cranked Slab Bar (45 deg)
        res4 = calculators.calculate_bbs_item('CRANK_SLAB', 10.0, {'a': 5.0, 'h': 0.12, 'hook': 0.15}, 6, 10)
        self.assertGreater(res4['cut_length_m'], 5.0)

        # 5. Rectangular Stirrup (135 deg seismic hooks)
        res5 = calculators.calculate_bbs_item('RECT_STIRRUP', 8.0, {'a': 0.25, 'b': 0.45}, 5, 20)
        self.assertGreater(res5['cut_length_m'], 1.4)

        # 6. Circular Tie
        res6 = calculators.calculate_bbs_item('CIRCULAR_TIE', 10.0, {'a': 0.60}, 2, 15)
        self.assertGreater(res6['cut_length_m'], 1.9)

        # 7. Chair Bar
        res7 = calculators.calculate_bbs_item('CHAIR_BAR', 12.0, {'a': 0.35, 'b': 0.15, 'c': 0.20}, 10, 5)
        self.assertAlmostEqual(res7['cut_length_m'], 1.05, places=2)

    def test_concrete_mix_grades(self):
        for grade in ['M10', 'M15', 'M20', 'M25', 'M30', 'M35']:
            res = calculators.calculate_concrete_materials(5.0, grade, 2.0)
            self.assertGreater(res['cement']['bags'], 0)
            self.assertGreater(res['fine_aggregate_sand']['weight_tons'], 0)
            self.assertGreater(res['coarse_aggregate']['weight_tons'], 0)

    def test_brickwork_and_plaster_calculators(self):
        # Brickwork
        res_b = calculators.calculate_brickwork(15.0, 3.2, 0.23, 'MODULAR', '1:6', 2.0, 5.0)
        self.assertGreater(res_b['total_bricks_count'], 4000)
        self.assertGreater(res_b['cement_bags'], 5.0)

        # Plastering
        res_p = calculators.calculate_plastering(200.0, 12.0, '1:4', 15.0)
        self.assertGreater(res_p['cement_bags'], 10.0)
        self.assertGreater(res_p['sand_tons'], 2.0)

        # Earthwork
        res_e = calculators.calculate_earthwork(25.0, 18.0, 3.5, 0.5)
        self.assertGreater(res_e['in_situ_volume_m3'], 1500.0)
        self.assertGreater(res_e['truck_loads_10m3'], 150)

    def test_rate_analysis_breakdown(self):
        mats = [{'name': 'Cement', 'qty': 8.0, 'rate': 7.50, 'waste_pct': 2.0}]
        labs = [{'name': 'Mason', 'qty': 0.5, 'rate': 40.00}]
        eqs = [{'name': 'Mixer', 'qty': 0.1, 'rate': 50.00}]
        res = calculators.compute_rate_analysis(mats, labs, eqs, 0.0, 1.5, 10.0, 5.0, 1.0, 1.0, 'm3')
        self.assertGreater(res['calculated_unit_rate'], 80.0)

    def test_evm_metrics(self):
        metrics = calculators.calculate_evm_metrics(100000.0, 110000.0, 100000.0, 500000.0)
        self.assertEqual(metrics['cost_variance_cv'], 10000.0)
        self.assertEqual(metrics['schedule_variance_sv'], 10000.0)
        self.assertEqual(metrics['cpi'], 1.1)
        self.assertEqual(metrics['spi'], 1.1)
        self.assertEqual(metrics['cost_status'], "On Budget")

    def test_rest_api_lifecycle(self):
        # 1. Projects
        res = self.app.get('/api/projects')
        self.assertEqual(res.status_code, 200)
        data = json.loads(res.data)
        self.assertTrue(data['success'])
        proj_id = data['projects'][0]['id']

        # 2. Add and Update BOQ item
        new_boq = {
            'item_code': '09.99',
            'division': '08. EXTERNAL WORKS & INFRASTRUCTURE',
            'section': 'Special Water Fountain',
            'description': 'Construct architectural cascade water fountain with LED lighting',
            'unit': 'ls',
            'quantity': 1.0,
            'unit_rate': 15000.00
        }
        add_res = self.app.post(f'/api/projects/{proj_id}/boq', json=new_boq)
        self.assertEqual(add_res.status_code, 201)
        item_id = json.loads(add_res.data)['item_id']

        # Update it
        new_boq['quantity'] = 2.0
        upd_res = self.app.put(f'/api/boq/{item_id}', json=new_boq)
        self.assertEqual(upd_res.status_code, 200)

        # Delete it
        del_res = self.app.delete(f'/api/boq/{item_id}')
        self.assertEqual(del_res.status_code, 200)

        # 3. Add QTO and Sync to BOQ
        qto_payload = {
            'name': 'Test Courtyard Area',
            'category': 'Finishes',
            'tool_type': 'polygon_area',
            'measured_value': 120.0,
            'unit': 'm2',
            'multiplier': 1.0,
            'depth_height': 0.0,
            'deduction_value': 10.0,
            'linked_boq_id': 'boq-16'
        }
        qto_res = self.app.post(f'/api/projects/{proj_id}/qto', json=qto_payload)
        self.assertEqual(qto_res.status_code, 201)
        qto_id = json.loads(qto_res.data)['takeoff_id']
        self.assertEqual(json.loads(qto_res.data)['net_quantity'], 110.0)

        # Clean up QTO
        self.app.delete(f'/api/qto/{qto_id}')

        # 4. Add IPC Certificate
        ipc_payload = {
            'cert_number': 99,
            'period_start': '2026-08-01',
            'period_end': '2026-08-31',
            'submission_date': '2026-09-02',
            'gross_work_this_period': 100000.0,
            'notes': 'Test certificate'
        }
        ipc_res = self.app.post(f'/api/projects/{proj_id}/ipc', json=ipc_payload)
        self.assertEqual(ipc_res.status_code, 201)
        # Gross 100k - Retention 5k - Advance 10k - Tax 5k = 80k
        self.assertEqual(json.loads(ipc_res.data)['net_payable'], 80000.0)

    def test_excel_and_csv_exports(self):
        projects = db.get_all_projects()
        proj = projects[0]
        proj_id = proj['id']

        # BOQ Excel
        excel_boq_res = self.app.get(f'/api/export/excel/boq/{proj_id}')
        self.assertEqual(excel_boq_res.status_code, 200)
        wb_boq = openpyxl.load_workbook(io.BytesIO(excel_boq_res.data))
        self.assertIn("Executive Summary", wb_boq.sheetnames)
        self.assertIn("Detailed BOQ", wb_boq.sheetnames)

        # BBS Excel
        excel_bbs_res = self.app.get(f'/api/export/excel/bbs/{proj_id}')
        self.assertEqual(excel_bbs_res.status_code, 200)
        wb_bbs = openpyxl.load_workbook(io.BytesIO(excel_bbs_res.data))
        self.assertIn("Bar Bending Schedule", wb_bbs.sheetnames)

        # BOQ CSV
        csv_res = self.app.get(f'/api/export/csv/boq/{proj_id}')
        self.assertEqual(csv_res.status_code, 200)
        self.assertIn(b"Item Code,Division,Section", csv_res.data)

if __name__ == '__main__':
    unittest.main()
